"""
Created on Feb. 19th  2022
@author: Raul Lopez-Lozano, Jingwen Wang
---------------------------------------------------------------------------------------------------------------------------
This library implements creates an HMC inversion Experiment that sample the posterior distributions of RTM model variables.
---------------------------------------------------------------------------------------------------------------------------
"""
from decimal import *
from openpyxl.reader.excel import load_workbook
from scipy.sparse.linalg import splu
getcontext().prec=100
import numpy as np
import pandas as pd
import scipy as sp
import shelve
from datetime import datetime
import math
from scipy import sparse


#Set of preliminary functions

def ArrayfromRanges(ranges,container):
    #Function to fill array container (empty) with ranges from an Excel sheet
    r=0
    for row in ranges:
        c=0
        for cell in row:
            cur_val=cell.value
            container[r,c]=cur_val
            c=c+1
        r=r+1
    return container

def ParameterDataFrame(sheet,nrows):
        previous_level=-1
        param_name=[]
        for row in sheet['B2':'B'+str(nrows)]:
            param_name.append(row[0].value)
        attribute_name=[]
        for col in sheet['C1':'K1']:
            for cell in col:
                attribute_name.append(cell.value)
        attribute_name.append('number_in_level')
        attribute_name.append('model')
        DF=pd.DataFrame([],index=param_name,columns=attribute_name)
        r=0
        for row in sheet['C2':'K'+str(nrows)]:
            c=0
            for cell in row:
                DF.iloc[r,c]=cell.value
                c=c+1
            r=r+1
        groups=np.zeros_like(np.unique(DF.iloc[:,4]), dtype=int)
        r=0
        for row in DF.index.values:
            DF.iloc[r,DF.columns.get_loc('number_in_level')]=groups[DF.iloc[r,DF.columns.get_loc('level')]]
            groups[DF.iloc[r,DF.columns.get_loc('level')]]=groups[DF.iloc[r,DF.columns.get_loc('level')]]+1
            r=r+1
        r=0
        for row in sheet['A2':'A'+str(nrows)]:
            DF.iloc[r,DF.columns.get_loc('model')]=row[0].value
            r=r+1
        return DF


def WriteinLogFile(citer,param, old_posterior,new_posterior,rmse_ref):
    f=open('debug_file.txt','a')
    f.write('iteration %u; parameter %s; old posterior %f; new posterior %f; RMSE %f\n' % (citer,param,old_posterior,new_posterior,rmse_ref) )
    f.close()

def getSoilSpeccoef(wvlength):
    import numpy as np
    rlambda=np.arange(400,2501,5)
    phis10=np.array([0.081,0.089,0.098,0.107,0.116,0.125,0.133,0.142,0.151,0.159,0.168,0.176,0.185,0.194,0.202,0.211,0.219,0.228,0.236,0.245,0.254,0.262,0.271,0.280,0.289,0.298,0.308,0.318,0.328,0.338,0.349,0.361,0.373,0.385,0.398,0.410,0.421,0.431,0.441,0.450,0.459,0.466,0.474,0.482,0.489,0.497,0.504,0.512,0.519,0.527,0.534,0.542,0.549,0.557,0.565,0.573,0.581,0.589,0.598,0.606,0.615,0.624,0.633,0.642,0.651,0.661,0.670,0.679,0.688,0.697,0.706,0.714,0.723,0.731,0.739,0.747,0.754,0.762,0.769,0.776,0.782,0.789,0.794,0.800,0.805,0.810,0.814,0.819,0.822,0.826,0.829,0.833,0.836,0.840,0.844,0.847,0.851,0.855,0.859,0.863,0.867,0.871,0.876,0.880,0.884,0.889,0.894,0.898,0.903,0.908,0.913,0.918,0.924,0.929,0.935,0.940,0.946,0.952,0.958,0.963,0.969,0.975,0.980,0.986,0.992,0.997,1.003,1.008,1.014,1.019,1.025,1.030,1.035,1.041,1.046,1.051,1.056,1.061,1.065,1.069,1.074,1.078,1.081,1.085,1.088,1.092,1.095,1.098,1.100,1.103,1.106,1.108,1.111,1.114,1.116,1.119,1.122,1.125,1.128,1.131,1.133,1.136,1.139,1.142,1.146,1.149,1.153,1.156,1.160,1.164,1.168,1.173,1.177,1.181,1.184,1.187,1.189,1.191,1.191,1.191,1.191,1.190,1.188,1.186,1.183,1.179,1.174,1.170,1.164,1.158,1.151,1.143,1.136,1.127,1.117,1.101,1.081,1.055,1.020,0.979,0.935,0.889,0.845,0.811,0.788,0.773,0.762,0.754,0.748,0.744,0.743,0.743,0.746,0.750,0.757,0.766,0.777,0.789,0.803,0.817,0.832,0.848,0.864,0.881,0.898,0.914,0.930,0.945,0.960,0.974,0.988,1.001,1.014,1.026,1.037,1.048,1.058,1.068,1.077,1.085,1.093,1.100,1.107,1.113,1.119,1.125,1.130,1.135,1.139,1.143,1.147,1.150,1.153,1.155,1.157,1.159,1.161,1.162,1.162,1.163,1.163,1.162,1.161,1.159,1.157,1.154,1.151,1.148,1.144,1.139,1.134,1.129,1.122,1.116,1.109,1.101,1.094,1.088,1.082,1.078,1.074,1.071,1.066,1.061,1.056,1.050,1.044,1.038,1.030,1.019,1.004,0.985,0.962,0.934,0.895,0.843,0.780,0.707,0.625,0.551,0.495,0.456,0.422,0.389,0.360,0.338,0.326,0.323,0.323,0.325,0.329,0.334,0.340,0.347,0.355,0.364,0.374,0.385,0.397,0.408,0.420,0.433,0.446,0.459,0.472,0.486,0.500,0.514,0.528,0.542,0.555,0.569,0.582,0.595,0.608,0.621,0.633,0.646,0.659,0.671,0.683,0.695,0.707,0.719,0.731,0.741,0.750,0.759,0.766,0.773,0.778,0.783,0.786,0.789,0.791,0.793,0.795,0.797,0.798,0.799,0.800,0.801,0.801,0.801,0.801,0.801,0.800,0.799,0.798,0.796,0.794,0.792,0.789,0.785,0.782,0.778,0.773,0.768,0.763,0.757,0.751,0.745,0.738,0.731,0.723,0.715,0.707,0.698,0.688,0.678,0.668,0.657,0.645,0.632,0.620,0.606,0.592,0.577,0.562,0.546,0.530,0.515,0.499,0.484,0.469,0.454,0.439,0.424,0.410,0.396,0.381,0.367,0.353,0.339,0.324,0.310,0.295,0.280,0.265,0.250,0.235])
    phis20=np.array([0.253,0.248,0.243,0.238,0.232,0.226,0.219,0.212,0.205,0.197,0.189,0.180,0.171,0.161,0.151,0.141,0.130,0.119,0.107,0.095,0.082,0.069,0.055,0.041,0.026,0.011,-.005,-.021,-.038,-.055,-.073,-.091,-.109,-.127,-.146,-.164,-.183,-.201,-.220,-.239,-.256,-.271,-.284,-.296,-.305,-.313,-.320,-.324,-.327,-.329,-.332,-.335,-.338,-.341,-.344,-.348,-.351,-.355,-.359,-.363,-.367,-.371,-.374,-.377,-.380,-.381,-.383,-.383,-.383,-.383,-.381,-.380,-.377,-.374,-.371,-.367,-.362,-.357,-.351,-.344,-.337,-.330,-.322,-.315,-.307,-.300,-.292,-.285,-.278,-.270,-.263,-.255,-.248,-.241,-.233,-.226,-.218,-.211,-.204,-.196,-.189,-.182,-.175,-.168,-.161,-.154,-.147,-.140,-.133,-.126,-.119,-.113,-.106,-.099,-.093,-.086,-.080,-.073,-.067,-.060,-.053,-.046,-.039,-.032,-.024,-.017,-.009,-.001,0.007,0.015,0.023,0.032,0.040,0.049,0.058,0.067,0.076,0.085,0.095,0.104,0.114,0.124,0.134,0.144,0.154,0.164,0.175,0.185,0.195,0.206,0.216,0.227,0.237,0.247,0.258,0.268,0.279,0.289,0.300,0.310,0.321,0.332,0.342,0.353,0.363,0.374,0.385,0.395,0.406,0.417,0.428,0.438,0.449,0.460,0.471,0.483,0.494,0.506,0.518,0.531,0.544,0.556,0.570,0.583,0.597,0.611,0.625,0.639,0.654,0.669,0.684,0.699,0.715,0.730,0.745,0.759,0.772,0.784,0.796,0.807,0.817,0.826,0.834,0.842,0.849,0.855,0.860,0.865,0.868,0.871,0.873,0.876,0.879,0.882,0.885,0.889,0.893,0.898,0.903,0.908,0.914,0.920,0.926,0.933,0.940,0.948,0.956,0.964,0.973,0.982,0.990,0.999,1.007,1.014,1.022,1.029,1.036,1.042,1.048,1.054,1.060,1.065,1.070,1.074,1.079,1.083,1.086,1.090,1.093,1.095,1.098,1.100,1.103,1.105,1.108,1.111,1.114,1.117,1.120,1.123,1.127,1.130,1.134,1.137,1.141,1.145,1.149,1.153,1.157,1.161,1.165,1.169,1.172,1.175,1.177,1.178,1.179,1.180,1.180,1.180,1.179,1.178,1.176,1.174,1.171,1.169,1.166,1.163,1.160,1.157,1.153,1.146,1.133,1.115,1.089,1.052,1.002,0.940,0.874,0.807,0.740,0.677,0.624,0.582,0.550,0.527,0.510,0.500,0.497,0.501,0.507,0.514,0.523,0.533,0.544,0.557,0.570,0.584,0.598,0.613,0.629,0.646,0.663,0.681,0.700,0.719,0.738,0.757,0.775,0.792,0.809,0.826,0.842,0.857,0.872,0.887,0.901,0.914,0.927,0.940,0.952,0.963,0.974,0.985,0.995,1.004,1.013,1.022,1.030,1.037,1.044,1.051,1.057,1.063,1.068,1.073,1.077,1.081,1.084,1.087,1.089,1.091,1.092,1.093,1.093,1.093,1.092,1.091,1.089,1.087,1.085,1.082,1.078,1.074,1.070,1.065,1.059,1.053,1.047,1.040,1.033,1.025,1.017,1.009,0.999,0.990,0.980,0.970,0.959,0.947,0.935,0.923,0.910,0.897,0.883,0.869,0.855,0.840,0.824,0.808,0.792,0.775,0.757,0.740,0.721,0.703,0.683,0.664,0.644,0.623,0.602,0.581,0.559,0.536,0.513,0.490,0.466,0.442,0.417,0.392,0.366])
    phis30=np.array([-.455,-.412,-.369,-.327,-.286,-.245,-.204,-.164,-.124,-.085,-.046,-.008,0.030,0.067,0.104,0.140,0.176,0.211,0.246,0.280,0.314,0.348,0.381,0.413,0.445,0.477,0.507,0.538,0.567,0.597,0.625,0.653,0.681,0.708,0.735,0.761,0.785,0.809,0.831,0.852,0.871,0.889,0.906,0.921,0.935,0.948,0.959,0.969,0.978,0.986,0.994,1.000,1.006,1.011,1.014,1.017,1.019,1.020,1.020,1.019,1.018,1.015,1.012,1.008,1.004,0.998,0.992,0.985,0.978,0.969,0.960,0.950,0.940,0.929,0.916,0.904,0.890,0.876,0.861,0.845,0.828,0.811,0.794,0.775,0.756,0.736,0.716,0.695,0.673,0.650,0.627,0.605,0.583,0.561,0.541,0.520,0.501,0.481,0.463,0.444,0.427,0.409,0.392,0.375,0.358,0.341,0.324,0.307,0.290,0.273,0.257,0.240,0.224,0.207,0.191,0.174,0.158,0.142,0.125,0.109,0.093,0.076,0.060,0.043,0.027,0.011,-.006,-.023,-.039,-.056,-.072,-.089,-.105,-.120,-.135,-.150,-.165,-.179,-.192,-.205,-.218,-.231,-.243,-.255,-.266,-.277,-.288,-.298,-.308,-.318,-.328,-.337,-.346,-.355,-.363,-.372,-.379,-.387,-.394,-.402,-.409,-.417,-.425,-.433,-.442,-.450,-.459,-.469,-.478,-.488,-.498,-.507,-.516,-.525,-.533,-.540,-.548,-.555,-.561,-.567,-.572,-.576,-.578,-.579,-.578,-.576,-.573,-.568,-.562,-.555,-.546,-.537,-.527,-.516,-.503,-.484,-.456,-.420,-.375,-.323,-.264,-.202,-.138,-.086,-.051,-.030,-.015,-.005,-.001,0.001,0.002,0.001,-.001,-.004,-.010,-.018,-.028,-.040,-.054,-.070,-.089,-.108,-.128,-.147,-.167,-.186,-.205,-.225,-.244,-.263,-.281,-.299,-.316,-.333,-.349,-.365,-.380,-.394,-.408,-.422,-.435,-.447,-.459,-.470,-.481,-.490,-.499,-.507,-.515,-.522,-.528,-.533,-.538,-.542,-.545,-.547,-.549,-.550,-.550,-.550,-.549,-.547,-.544,-.541,-.537,-.531,-.526,-.519,-.511,-.503,-.494,-.485,-.476,-.468,-.460,-.452,-.445,-.439,-.433,-.427,-.422,-.418,-.416,-.415,-.415,-.416,-.413,-.406,-.395,-.381,-.363,-.343,-.321,-.289,-.231,-.147,-.071,-.013,0.032,0.077,0.122,0.160,0.185,0.199,0.203,0.205,0.207,0.208,0.209,0.209,0.210,0.210,0.209,0.209,0.208,0.206,0.204,0.202,0.200,0.197,0.194,0.191,0.187,0.183,0.178,0.174,0.168,0.163,0.157,0.150,0.143,0.136,0.129,0.121,0.112,0.104,0.095,0.086,0.077,0.068,0.060,0.051,0.043,0.034,0.026,0.018,0.009,0.001,-.007,-.015,-.022,-.030,-.037,-.043,-.049,-.054,-.059,-.063,-.067,-.071,-.074,-.076,-.078,-.079,-.080,-.081,-.081,-.080,-.079,-.077,-.075,-.073,-.070,-.067,-.064,-.060,-.056,-.052,-.047,-.042,-.037,-.031,-.025,-.019,-.012,-.005,0.003,0.010,0.019,0.027,0.036,0.045,0.055,0.065,0.076,0.087,0.098,0.110,0.122,0.134,0.144,0.153,0.160,0.165,0.168,0.170,0.171,0.169,0.166,0.162,0.157,0.151,0.144,0.137,0.128,0.119,0.109,0.098,0.086,0.074,0.061])
    phis40=np.array([0.058,0.070,0.081,0.092,0.103,0.112,0.122,0.131,0.139,0.148,0.155,0.162,0.169,0.175,0.181,0.186,0.191,0.196,0.200,0.203,0.206,0.209,0.211,0.213,0.214,0.215,0.215,0.215,0.214,0.213,0.212,0.210,0.207,0.204,0.201,0.197,0.192,0.187,0.182,0.176,0.169,0.162,0.154,0.146,0.138,0.129,0.119,0.109,0.099,0.089,0.079,0.069,0.059,0.049,0.039,0.029,0.019,0.009,-.001,-.011,-.021,-.031,-.041,-.051,-.060,-.069,-.077,-.085,-.092,-.099,-.105,-.111,-.116,-.121,-.125,-.129,-.133,-.136,-.138,-.140,-.142,-.143,-.145,-.147,-.148,-.150,-.151,-.153,-.154,-.156,-.157,-.159,-.160,-.162,-.163,-.164,-.165,-.165,-.165,-.165,-.165,-.164,-.163,-.162,-.161,-.159,-.157,-.155,-.152,-.149,-.146,-.142,-.138,-.133,-.128,-.122,-.116,-.109,-.102,-.094,-.085,-.076,-.067,-.057,-.046,-.035,-.024,-.012,0.000,0.012,0.024,0.036,0.049,0.061,0.074,0.087,0.100,0.113,0.126,0.140,0.153,0.167,0.181,0.195,0.210,0.224,0.238,0.253,0.267,0.282,0.297,0.311,0.326,0.341,0.356,0.371,0.385,0.400,0.415,0.430,0.446,0.461,0.477,0.493,0.510,0.527,0.544,0.562,0.579,0.597,0.614,0.631,0.648,0.664,0.680,0.695,0.710,0.725,0.738,0.752,0.764,0.776,0.787,0.797,0.805,0.811,0.815,0.816,0.816,0.816,0.818,0.821,0.820,0.811,0.793,0.765,0.720,0.658,0.579,0.494,0.408,0.322,0.242,0.181,0.141,0.119,0.108,0.106,0.113,0.124,0.134,0.142,0.149,0.157,0.169,0.184,0.204,0.228,0.255,0.283,0.313,0.344,0.377,0.411,0.446,0.480,0.514,0.547,0.579,0.610,0.640,0.669,0.698,0.726,0.752,0.777,0.801,0.824,0.846,0.867,0.886,0.904,0.921,0.937,0.952,0.967,0.980,0.992,1.004,1.015,1.025,1.034,1.042,1.050,1.056,1.062,1.067,1.071,1.074,1.077,1.078,1.079,1.077,1.074,1.069,1.063,1.054,1.044,1.033,1.019,1.004,0.988,0.971,0.954,0.937,0.921,0.906,0.893,0.882,0.872,0.865,0.858,0.851,0.844,0.838,0.830,0.823,0.814,0.805,0.790,0.761,0.716,0.655,0.579,0.491,0.394,0.287,0.171,0.045,-.078,-.177,-.252,-.303,-.330,-.333,-.331,-.329,-.327,-.326,-.324,-.323,-.321,-.319,-.316,-.313,-.309,-.304,-.298,-.293,-.288,-.282,-.277,-.271,-.265,-.257,-.247,-.236,-.222,-.207,-.190,-.172,-.154,-.137,-.122,-.107,-.093,-.080,-.068,-.055,-.042,-.028,-.013,0.004,0.021,0.039,0.058,0.078,0.095,0.109,0.119,0.127,0.130,0.131,0.128,0.121,0.112,0.102,0.092,0.084,0.075,0.067,0.060,0.053,0.047,0.042,0.037,0.032,0.028,0.025,0.021,0.018,0.015,0.011,0.008,0.004,0.001,-.003,-.006,-.010,-.014,-.017,-.021,-.025,-.028,-.032,-.036,-.040,-.044,-.047,-.051,-.054,-.056,-.059,-.060,-.062,-.063,-.063,-.063,-.063,-.063,-.062,-.060,-.058,-.056,-.053,-.050,-.047,-.045,-.042,-.040,-.039,-.037,-.035,-.034,-.033,-.033,-.032,-.032,-.032,-.032,-.033])
    phis1=np.interp(wvlength,rlambda,phis10)
    phis2=np.interp(wvlength,rlambda,phis20)
    phis3=np.interp(wvlength,rlambda,phis30)
    phis4=np.interp(wvlength,rlambda,phis40)
    coef = np.asarray((phis1,phis2,phis3,phis4)).transpose()
    return coef


def skylspec(wvl,bangs):
    import numpy as np
    in_wvl=np.arange(400,2501,5)
    gamml0=np.append(np.array([.426, .438, .450, .461, .472, .483, .493, .503, .513, .523,.532, .541, .550, .559, .567, .575, .583, .591, .598, .606,.613, .620, .626, .633, .640, .647, .654, .661, .668, .675,.683, .690, .697, .705, .712, .719, .726, .733, .739, .744,.748, .750, .752, .752, .751, .749, .746, .743, .740, .736,.732, .728, .724, .720, .716, .712, .707, .703, .698, .693,.689, .684, .679, .674, .668, .663, .657, .652, .646, .640,.634, .627, .617, .605, .590, .573, .553, .531, .505, .478,.448]),np.ones(340)*0.4)
    alpl0=np.append(np.array([.394, .381, .369, .357, .345, .334, .323, .312, .302, .293,.283, .275, .266, .258, .250, .243, .236, .230, .224, .218,.213, .208, .202, .197, .192, .187, .182, .176, .171, .166,.160, .155, .150, .144, .139, .133, .128, .122, .117, .112,.107, .102, .097, .092, .087, .083, .078, .074, .070, .066,.062, .058, .054, .050, .047, .043, .040, .037, .034, .031,.028, .025, .023, .020, .017, .015, .012, .010, .008, .006,.004, .003, .002, .001]), np.zeros(347))
    gamml=np.interp(wvl,in_wvl,gamml0)
    alpl=np.interp(wvl,in_wvl,alpl0)
    gammlc = -1.4
    tausk = np.power((wvl/1000),gammlc)*bangs
    skyl = tausk*gamml+alpl
    difuse_f= skyl/(1+skyl)
    direct_f=1-difuse_f
    return direct_f


#End of prelimnary functions


class Experiment():

    def __init__(self,excelWB):
        #Script to read the configuration of the experiment from an Excel Sheet and create from that the necessary variables to call HMC
        data_book=load_workbook(excelWB)
        ###########################################################################
        #First, reading general configurationof the experiment
        sheet_config=data_book.get_sheet_by_name('RunConfig')
        self.deltaGrad = 0.00001
        self.sigma_proposal=0.1
        self.r_proposal=0.1
        self.niter=sheet_config['B1'].value #Number of iterations for HMC
        self.nsensors=sheet_config['B2'].value #Number of sensors used in the analysis
        self.nlevels=sheet_config['B3'].value #Number of hierarchical levels of model variables (level 0: crop specific; level1: site specific)
        self.leaf_model=sheet_config['B4'].value
        self.soil_model=sheet_config['B5'].value
        self.skyl_model=sheet_config['B6'].value
        self.canopy_model=sheet_config['B7'].value
        start_epsilon=sheet_config['B8'].value
        self.bcNNs = shelve.open('emulators/'+sheet_config['B9'].value)
        self.sensor_names=[]
        self.sensor_nbands=[]
        self.sensor_band_names = []
        self.sensor_obs_ref=[]
        self.sensor_nobs=[]
        nobs=0
        ###########################################################################
        # Now, reading  observed reflectance from different sensors
        grouping=[]
        for i in np.arange(0,self.nsensors):
            self.sensor_names.append(sheet_config[chr(1+i+ord('A'))+'10'].value)
            sheet_obsref = data_book.get_sheet_by_name('obs_reflectance'+self.sensor_names[i])
            nrows = sheet_obsref.max_row
            ncolumns= sheet_obsref.max_column
            self.sensor_nobs.append(nrows - 1)
            nobs += nrows - 1
            self.sensor_nbands.append(int(ncolumns-2))
            self.sensor_band_names.append(ArrayfromRanges(sheet_obsref[chr(0 + ord('A')) + str(1): chr((0 + self.sensor_nbands[i] - 1) + ord('A')) + str(1)], np.chararray((1,self.sensor_nbands[i]), itemsize=6)))
            self.sensor_obs_ref.append(ArrayfromRanges(sheet_obsref[chr(0+ord('A'))+str(2) : chr((0+self.sensor_nbands[i]-1)+ord('A'))+str(nrows)],np.zeros((nrows-1,self.sensor_nbands[i]), dtype=float)))
            if i==0:
                grouping=ArrayfromRanges(sheet_obsref[chr(self.sensor_nbands[i]+ord('A'))+str(2) : chr((self.sensor_nbands[i] + self.nlevels-1)+ord('A'))+str(nrows)],np.zeros((nrows-1,self.nlevels)))
            else:
                grouping = np.vstack((grouping,ArrayfromRanges(sheet_obsref[chr(self.sensor_nbands[i]+ord('A'))+str(2) : chr((self.sensor_nbands[i] + self.nlevels-1)+ord('A'))+str(nrows)],np.zeros((nrows-1,self.nlevels)))))
        self.nobs=nobs
        self.hierarchical_groups=grouping
        sheet_obsref=None

        ##################################################################################
        #Now reading parameter list and parameter attributes
        sheet_param=data_book.get_sheet_by_name('paramvalues')
        nrows=sheet_param.max_row
        self.parameters_Frame=ParameterDataFrame(sheet_param, nrows) #Pandas DataFrame with variable parameter
        sheet_param=None
        #Now constructing array with future parameter values, sigma and r matrices
        self.old_posterior=0 #variable to store the previous posterior density calculated
        self.parameters_value=[]
        self.Minv = []
        self.Phi = []
        self.level_epsilon=[]
        for c_level in np.arange(0, self.nlevels):
            groups_level = np.unique(self.hierarchical_groups[:, c_level])
            number_groups_level = len(groups_level)
            n_param_level = np.sum(self.parameters_Frame['level'] == c_level)
            self.level_epsilon.append(np.zeros(number_groups_level))
            self.level_epsilon[c_level][:] = start_epsilon
            self.parameters_value.append(np.zeros((self.niter, n_param_level, number_groups_level)))
            self.Minv.append(np.zeros((n_param_level, number_groups_level)))
            self.Phi.append(np.zeros((n_param_level, number_groups_level)))
            SeriesParam = self.parameters_Frame.loc[self.parameters_Frame.level == c_level, 'number_in_level']
            for par in SeriesParam.index.values:
                start_val = self.parameters_Frame.loc[par, 'start_value']
                Minv = self.parameters_Frame.loc[par, 'Minv_HMC']
                Phi = self.parameters_Frame.loc[par, 'Phi_HMC']
                self.Minv[c_level][self.parameters_Frame.loc[par, 'number_in_level'], :] = Minv
                self.Phi[c_level][self.parameters_Frame.loc[par, 'number_in_level'], :] = Phi
                if start_val == 'random':
                    self.parameters_value[c_level][:, self.parameters_Frame.loc[par, 'number_in_level'], :] = np.tile(
                        np.random.uniform(self.parameters_Frame.loc[par, 'min_value'],
                                          self.parameters_Frame.loc[par, 'max_value'], (1, number_groups_level)),
                        (self.niter, 1))
                elif start_val == 'fixed':
                    sheet_start = data_book.get_sheet_by_name('obs_starting_values')
                    ncolsb = sheet_start.max_column
                    nrowsb = sheet_start.max_row
                    for r in sheet_start['A1':chr(ncolsb - 1 + ord('A')) + str(1)]:
                        c_col = 0
                        for col in r:
                            if col.value == par:
                                c_rec = 0
                                for record in sheet_start[
                                              chr(c_col + ord('A')) + '2':chr(c_col + ord('A')) + str(nrowsb)]:
                                    self.parameters_value[c_level][:, self.parameters_Frame.loc[par, 'number_in_level'],
                                    c_rec] = record[0].value
                                    c_rec = c_rec + 1
                                del (c_rec)
                            c_col = c_col + 1
                    del (sheet_start)
                    del (ncolsb)
                    del (nrowsb)
                    del (c_col)
                    del (r)
                else:
                    self.parameters_value[c_level][:, self.parameters_Frame.loc[par, 'number_in_level'], :] = start_val
        ##################################################################################
        #Now reading parameters with prior distributions per observations (if needed)
        if np.sum(self.parameters_Frame['prior_dist_a']=='obs_prior')>0:
            sheet_priors=data_book.get_sheet_by_name('obs_prior')
            nrows=sheet_priors.max_row
            ncols=sheet_priors.max_column
            n_param=[]
            for c_col in sheet_priors['A1':chr(ncols-1 + ord('A'))+str(1)]:
                n_param.append(c_col[0].value)
            self.parameters_prior_values=pd.DataFrame([],index=np.arange(0,nrows-1),columns=n_param)
            c=0
            ranges_prior=sheet_priors['A2':chr(ncols-1 + ord('A'))+str(nrows)]
            for c_col in ranges_prior:
                r=0
                for cell in c_col:
                    self.parameters_prior_values.iloc[c,r]=cell.value
                    r=r+1
                c=c+1
        else:
            self.parameters_prior_values=None
        sheet_priors=None
        ########################################################################################
        #Now reading fixed parameters
        sheet_fixed=data_book.get_sheet_by_name('fixedparameters')
        nrows=sheet_fixed.max_row
        ncols=sheet_fixed.max_column
        fixed_values=np.zeros((nrows-2,ncols))
        ranges_values = sheet_fixed['A3':chr(ncols-1+ord('A'))+str(nrows)]
        fixed_values=ArrayfromRanges(ranges_values,fixed_values)
        ranges_model = sheet_fixed['A1':chr(ncols - 1 + ord('A')) + str(1)]
        ranges_param = sheet_fixed['A2':chr(ncols - 1 + ord('A')) + str(2)]
        nl=0
        for rows in ranges_param:
            c=0
            for cell in rows:
                self.parameters_Frame.loc[cell.value,'level']=self.nlevels
                self.parameters_Frame.loc[cell.value, 'model'] = ranges_model[0][0].value
                self.parameters_Frame.loc[cell.value,'number_in_level']=nl
                nl=nl+1
                c=c+1
        self.parameters_value.append(fixed_values)
        ##################################################################################
        self.RTMModelSetUp=RTMmodelSolution(self) #RTM Model solution : leaf model + soil model + direct solar beam model + canopy model
        ##################################################################################
        print(self.parameters_Frame)


    def runExperiment (self,siter):
    #This function runs the HMC created, from the staring iteration siter. For a new experiment, siter=0 (starting from the first iteration)
        target_acc_rate=0.75
        n_acc=[]
        for level in np.arange(0,self.nlevels):
            n_acc.append(np.zeros(len(np.unique(self.hierarchical_groups[:,level]))))
        ############################
        ##Now starting interations##
        ############################
        for citer in np.arange(siter+1,self.niter):
            #Update parameters value from previous iteration
            self.UpdateNewIter(citer)
            for clevel in np.arange(0,self.nlevels):
                n_parallel = len(np.unique(self.hierarchical_groups[:,clevel]))
                n_param = int(np.sum(self.parameters_Frame['level'].values==clevel))
                Phi = np.tile(np.transpose(1 / np.asarray(self.parameters_Frame['Phi_HMC'].values[ self.parameters_Frame['level']==clevel],dtype=float)),(n_parallel,1)) * np.asarray(np.random.normal(0,1,(n_parallel, n_param)))
                M_inv = np.tile(np.transpose(np.power(np.asarray(self.parameters_Frame['Minv_HMC'].values[ self.parameters_Frame['level']==clevel],dtype=float),2)),(n_parallel,1))
                #np.asarray(np.reshape(np.power(self.parameters_Frame['Minv_HMC'].values[ self.parameters_Frame['level']==clevel],2),(-1,1)),dtype=float)
                #####################################################################################################
                parameters_value = []
                for i in np.arange(0, self.nlevels):
                    parameters_value.append(self.parameters_value[i][citer, :, :].copy())
                parameters_value.append(self.parameters_value[self.nlevels].copy())
                reflectance_bands_orig = self.RTMModelSetUp.runModelSolution(parameters_value, self.nlevels, self)
                PosteriorOld=self.ComputeLogPosterior(reflectance_bands_orig,clevel,parameters_value)
                HMCScore_old = PosteriorOld - 0.5 * np.reshape(np.sum(M_inv* np.power(Phi, 2),axis=1),(-1,1))
                reflectance_bands = self.RTMModelSetUp.runModelSolution(parameters_value, clevel, self)
                ParGradients = self.ComputeGradients(reflectance_bands, clevel, parameters_value)
                iter_epsilon = np.tile(np.reshape(np.random.uniform(0, self.level_epsilon[clevel][:] * 2),(-1,1)),(1,n_param))
                iter_leapfrog_steps = np.ceil(np.random.uniform(0, 10))
                for leapfrog_step in np.arange(0, iter_leapfrog_steps):
                    Phi += 0.5 * iter_epsilon * ParGradients  # Update phi
                    leapfrog_iter_epsilon = iter_epsilon.copy()
                    while np.sum(leapfrog_iter_epsilon) > 0:
                        epsilon_to_closest_boundary = self.PredictCollision(Phi, M_inv, parameters_value, clevel)
                        hitting_boundaries = np.less(epsilon_to_closest_boundary , leapfrog_iter_epsilon)
                        temp_epsilon = np.minimum(leapfrog_iter_epsilon, epsilon_to_closest_boundary)
                        delta_parameters = np.asarray(temp_epsilon * M_inv * Phi)
                        parameters_value[clevel] += np.transpose(delta_parameters)
                        Phi[hitting_boundaries] = Phi[hitting_boundaries] *-1 #Reflection
                        leapfrog_iter_epsilon -= temp_epsilon
                    reflectance_bands = self.RTMModelSetUp.runModelSolution(parameters_value, clevel, self)
                    ParGradients = self.ComputeGradients(reflectance_bands, clevel, parameters_value)
                    error = np.isnan(ParGradients)
                    if error.any() == True:
                        print('Warning: nan in gradients')
                    Phi += 0.5 * iter_epsilon * ParGradients
                Phi = Phi * -1
                reflectance_bands_orig = self.RTMModelSetUp.runModelSolution(parameters_value, self.nlevels, self)
                PosteriorNew = self.ComputeLogPosterior(reflectance_bands_orig, clevel, parameters_value)
                HMCScore_new = PosteriorNew - 0.5 * np.reshape(np.sum(M_inv * np.power(Phi, 2), axis=1), (-1, 1))
                update = (HMCScore_new - HMCScore_old) > np.log(np.random.uniform(0, 1, 1))
                if np.any(update):
                    self.parameters_value[clevel][citer,:,update[:,0]] = np.transpose(parameters_value[clevel][:,update[:,0]])
                    n_acc[clevel][update[:,0]]+=1
                if clevel == 0:
                    testHMC_old=PosteriorOld
                    testHMC_new=PosteriorNew

            parameters_value = []
            for i in np.arange(0, self.nlevels):
                parameters_value.append(self.parameters_value[i][citer, :, :].copy())
            parameters_value.append(self.parameters_value[self.nlevels].copy())
            reflectance_bands = self.RTMModelSetUp.runModelSolution(parameters_value, self.nlevels, self)
            SSE=0
            den=0
            for sensor in np.arange(0,self.nsensors):
                    SSE += np.sum(np.power(np.transpose(self.sensor_obs_ref[sensor]) - reflectance_bands[sensor],2).flatten())
                    den += np.shape(reflectance_bands[sensor])[0] * np.shape(reflectance_bands[sensor])[1]
            RMSE=np.sqrt(SSE/den)
            param=self.parameters_value[0][citer,:,0]
            print('scores ' + str(citer) + ' ' + str(testHMC_old) + ' ' +str(testHMC_new))
            print('Param ' + str(citer) + ' ' + str(param))
            print('RMSE reflectance = ' + str(np.round(RMSE,5)))
            #for debugging in HPC
            f = open('logfileHMC1.txt', 'a')
            f.write('Time %s; iteration %u; posterior %f; HMC score new %f; param values %f %f %f %f %f; RMSE %f\n' % ( str(datetime.now()),citer, testHMC_old, testHMC_new, param[0], param[1],param[2],param[3],param[4], RMSE))
            f.close()
            #Comment if run outside
            if (citer + 1) % 200 == 0:
                print("Aceptance rate level 0 = " + str(np.round(n_acc[0] / 200, 2)))
                print("Aceptance rate level 1 = " + str(np.round(np.mean(n_acc[1] / 200),2)))
                # if citer + 1 == 500:
                for clevel in np.arange(0,self.nlevels):
                    self.level_epsilon[clevel][:]=self.level_epsilon[clevel][:]*((n_acc[clevel]/200)/target_acc_rate)
                print("epsilon level 0  = " + str(np.round(np.mean(self.level_epsilon[0][:]), 5)))
                n_acc[0][:]=0
                n_acc[1][:]=0
            #Finish iteration
            ####################################################################################################################

    def ComputeLogLikelihood(self,reflectance, level, parameters_value):
        # This function computes the log Likelihood for a set of reflectance observations (multiple sensors) given the simulated reflectance
        if level < self.nlevels:
            nsims_per_obs = np.sum(self.parameters_Frame['level'].values==level) * 2
            npar=np.sum(self.parameters_Frame['level'].values==level)
        else:
            nsims_per_obs=1
            npar=1
        logLikelihood=np.zeros(self.nobs*nsims_per_obs)
        offset_sigma=np.zeros((nsims_per_obs,1))
        if self.parameters_Frame.loc['sigma','level']==level:
            pos_sigma=self.parameters_Frame.loc['sigma','number_in_level']
            offset_sigma[pos_sigma]=self.deltaGrad
            offset_sigma[pos_sigma+npar]= -1*self.deltaGrad
        position=0
        for sensor in np.arange(0,len(self.sensor_names)):
            sigma = np.repeat(np.reshape(parameters_value[self.parameters_Frame.loc['sigma', 'level']][
                                         self.parameters_Frame.loc['sigma', 'number_in_level'], position: position + self.sensor_nobs[sensor]], (-1, 1)),
                              nsims_per_obs, axis=0) + np.tile(offset_sigma, (self.sensor_nobs[sensor], 1))
            error_vector = np.repeat(self.sensor_obs_ref[sensor],nsims_per_obs, axis=0) - np.transpose(reflectance[sensor])
            n = np.shape(error_vector)[1]
            SSE=np.reshape(np.asarray(np.sum(np.power(error_vector, 2),axis=1)),(-1,1))
            logLikelihood[position * nsims_per_obs: position * nsims_per_obs + self.sensor_nobs[sensor]*nsims_per_obs] = (-n * np.log(sigma) - 1 / (2 * sigma * sigma) * SSE)[:,0]
            position+=self.sensor_nobs[sensor]
        return logLikelihood

    def ComputeLogPriorProbability(self, size_sim,level, parameters_value, n_param):
        # This function computes the log Prior probability for a set of RTM variables
        parameters_prior=np.logical_and(np.logical_and(self.parameters_Frame['level']== level , self.parameters_Frame['prior_dist']!='uniform'),self.parameters_Frame['level'].values< self.nlevels)
        if n_param==0:
            repeat_l=1
        else:
            repeat_l=n_param* 2
        log_prior_parameter = np.zeros((size_sim*repeat_l,1))
        if np.any(parameters_prior):
            parameters_level_prior=self.parameters_Frame.loc[np.logical_and(self.parameters_Frame['level']== level , self.parameters_Frame['prior_dist']!='uniform')]
            if n_param == 0:
                offset=np.zeros((1,np.sum(self.parameters_Frame['level'].values== level)))
            else:
                offset=np.tile(np.vstack((np.identity(n_param)*self.deltaGrad,np.identity(n_param)*self.deltaGrad*-1)),(np.shape(parameters_value[level])[1],1))
            for param in parameters_level_prior.index.values:
                values = np.reshape(np.transpose(
                    np.repeat(parameters_value[level][parameters_level_prior.loc[param, 'number_in_level'], :],
                              repeat_l)) + offset[:, parameters_level_prior.loc[param, 'number_in_level']], (-1, 1))
                if parameters_level_prior.loc[param, 'prior_dist'] == 'normal':
                    if parameters_level_prior.loc[param, 'prior_dist_a'] == 'obs_prior':
                        prior_mean = np.reshape(np.transpose(
                            np.repeat(np.asarray(self.parameters_prior_values[param].values, dtype=float), repeat_l)),
                                                (-1, 1))
                    else:
                        prior_mean = np.reshape(
                            np.repeat([float(parameters_level_prior.loc[param, 'prior_dist_a'])], len(values)), (-1, 1))
                    if parameters_level_prior.loc[param, 'prior_dist_b'] == 'obs_prior':
                        prior_std = np.reshape(np.transpose(
                            np.repeat(np.asarray(self.parameters_prior_values[param + '_b'].values, dtype=float),
                                      repeat_l)), (-1, 1))
                    else:
                        prior_std = np.reshape(
                            np.repeat([float(parameters_level_prior.loc[param, 'prior_dist_b'])], len(values)), (-1, 1))
                    log_prior_parameter += np.log((1 / (prior_std * np.sqrt(2 * np.pi))) * np.exp(
                        -1 * np.power(values - prior_mean, 2) / (2 * prior_std * prior_std)))
                elif parameters_level_prior.loc[param, 'prior_dist'] == 'n-i-inverse-gamma':
                    log_prior_parameter += np.log(1 / values)
                elif parameters_level_prior.loc[param, 'prior_dist'] == 'inverse-gamma':
                    log_prior_parameter += np.log(
                        sp.stats.invgamma.pdf(values, float(parameters_level_prior.loc[param, 'prior_dist_a']),
                                           loc=float(parameters_level_prior.loc[param, 'prior_dist_b']),
                                           scale=float(parameters_level_prior.loc[param, 'prior_dist_c'])))

        return log_prior_parameter

    def ComputeGradients(self,reflectance_modelled, level, parameters_value):
        #This function calculates the gradient of the log Posterior function for a given set of RTM variables
        LogLikelihood = self.ComputeLogLikelihood(reflectance_modelled,level,parameters_value)
        groups,counts=np.unique(self.hierarchical_groups[:,level],return_counts=True)
        mapping_out=np.repeat(np.identity(len(groups)),counts, axis=1)
        n_param = np.sum(self.parameters_Frame['level']== level)
        mapping_out_par = np.repeat(np.identity(len(groups)), n_param*2, axis=1)
        LogPrior = self.ComputeLogPriorProbability(len(groups),level,parameters_value,n_param)
        LogPosterior = np.dot(mapping_out, np.reshape(LogLikelihood, (self.nobs, n_param * 2))) + np.dot(mapping_out_par, LogPrior)
        Gradient = (LogPosterior[:,0:n_param] - LogPosterior[:,n_param:2*n_param])/ (self.deltaGrad*2)
        return Gradient

    def ComputeLogPosterior(self,reflectance_modelled, level, parameters_value):
        # This function calculates the log Posterior function for a given set of RTM variables
        LogLikelihood = self.ComputeLogLikelihood(reflectance_modelled,2,parameters_value)
        groups,counts=np.unique(self.hierarchical_groups[:,level],return_counts=True)
        mapping_out=np.repeat(np.identity(len(groups)),counts, axis=1)
        mapping_out_par = np.repeat(np.identity(len(groups)), 1, axis=1)
        LogPrior = self.ComputeLogPriorProbability(len(groups),level,parameters_value,0)
        LogPosterior=np.dot(mapping_out,np.reshape(LogLikelihood,(self.nobs,1))) + np.dot(mapping_out_par,np.reshape(LogPrior,(len(groups),1)))
        return LogPosterior

    def AddNewIterations (self):
        #This function starts a new Experiment from the last iteration of an existing one, when more iterations are needed to converge
        for c_level in np.arange(0,self.nlevels):
            curr_param=self.parameters_value[c_level][self.niter-1,:,:]
            new_param=np.zeros_like(self.parameters_value[c_level])
            new_param[:]=curr_param
            self.parameters_value[c_level][:]=new_param[:]
            del(new_param)


    def UpdateNewIter(self,citer):
        #Updates the iteration when th previous one has been accepted by the HMC sampler
        for level in np.arange(0,self.nlevels):
            for group in np.unique(self.hierarchical_groups[:,level]):
                self.parameters_value[level][citer,:,int(group)]=self.parameters_value[level][citer-1,:,int(group)]

    def PredictCollision(self,Phi,M_inv,param_values_all,level):
        param_values=param_values_all[level]
        min_val = np.tile(np.reshape(np.asarray(self.parameters_Frame['min_value'].values[self.parameters_Frame['level'].values==level], dtype=float),(-1,1)),(1,np.shape(param_values)[1])) #,np.shape(param_values)[1])
        max_val = np.tile(np.reshape(np.asarray(self.parameters_Frame['max_value'].values[self.parameters_Frame['level'].values==level], dtype=float),(-1,1)),(1,np.shape(param_values)[1])) #,np.shape(param_values)[1])
        delta_tomax = np.transpose(max_val - param_values)
        delta_tomin = np.transpose(min_val - param_values)
        epsilon_to_boundaries=np.zeros((np.shape(Phi)[0],np.shape(Phi)[1],2))
        epsilon_to_boundaries[:,:,0]= delta_tomax/ (M_inv * Phi)
        epsilon_to_boundaries[:,:,1]= delta_tomin/ (M_inv * Phi)
        epsilon_to_closest_boundary=np.max(epsilon_to_boundaries,axis=2)
        boundary_values = np.transpose(param_values) + epsilon_to_closest_boundary * Phi * M_inv #for debugging, then remove
        return epsilon_to_closest_boundary






class RTMmodelSolution():
    #A model solution is a suite of RTM models that are linked to model canopy reflectance, including:
    #   1) A leaf optical properties model emulator
    #   2) A soil reflectance model (embedded, from Price 1990)
    #   3) A model that simulates the fraction of direct beam (embedded)
    #   4) A canopy reflectance model emulator
    #   5) A Neural Network for bands convolution
    #   The sequence of model inputs and outputs have to be hardcoded in this section
    #   Both the leaf optical properties and soil reflectance models are object of the SpectralModels class
    #   The canopy reflectance model is an object of the CanopyModel class
    def __init__(self, Experiment):
        #creating the model solution class from the information contained in the Experiment
        self.SpectralModels=[]
        self.nlevels=Experiment.nlevels
        sensors=Experiment.sensor_names
        band_names=Experiment.sensor_band_names
        nobs=Experiment.nobs
        self.sensor_number = np.zeros(nobs, dtype=int)
        ####################################################
        myshelf = shelve.open('emulators/'+Experiment.canopy_model)
        self.wvl_emulation = myshelf['selected_Bands']  ###### New
        myshelf.close()
        ####################################################
        self.CanopyModel=CanopyModel(Experiment)
        self.Fdirect_values=skylspec(self.wvl_emulation,0.19).reshape((-1,1)) ############# New #############
        self.Fdirect_param_position=np.where(self.CanopyModel.sequence=='fdir')[0]
        #Setting up Spectral models (leaf reflectance model and soil reflectance model)
        if Experiment.leaf_model == 'emulator_PROSPECT40':
            sequence_leaf = ['N','Cab','Car','Cw','Cm'] #Order of parameters in PROSPECT 5 emulator
            coef_leaf= None  #
        if Experiment.soil_model == 'soilspec':
            sequence_soil= ['rsl1','rsl2','rsl3','rsl4'] #Order of parameters in soilspec
            coef_soil=getSoilSpeccoef(self.wvl_emulation)
        self.SpectralModels.append(SpectralModel(Experiment.leaf_model,sequence_leaf,['leafr', 'leaft'],self.CanopyModel.sequence,Experiment, self.wvl_emulation, coef_leaf)) ############# New #############
        self.SpectralModels.append(SpectralModel(Experiment.soil_model, sequence_soil, ['soilr'], self.CanopyModel.sequence,Experiment,self.wvl_emulation, coef_soil)) ############# New #############

    def runModelSolution(self,parameter_values,level,Experiment):
        model_input = self.CanopyModel.prepareInput(level,parameter_values,Experiment)
        for SpModel in self.SpectralModels:
            model_input = SpModel.runSpectralModel(level,model_input,parameter_values)
        model_input[:,self.Fdirect_param_position]=np.tile(self.Fdirect_values,(int(len(model_input)/len(self.Fdirect_values)),1))
        reflectance=self.CanopyModel.Run(model_input)
        nsims_per_obs = int(len(reflectance)/(Experiment.nobs*len(self.wvl_emulation))) ##########New############
        reflectance_bands=[]
        init=0
        for i in np.arange(0,len(Experiment.sensor_names)):
            sensor_name = Experiment.sensor_names[i]
            rs_ref = np.reshape(
                reflectance[init:init + nsims_per_obs * len(self.wvl_emulation) * Experiment.sensor_nobs[i]],
                (len(self.wvl_emulation), nsims_per_obs * Experiment.sensor_nobs[i]),
                order='F')
            reflectance_bands.append((Experiment.bcNNs[sensor_name+"_SC_y"].inverse_transform(
                Experiment.bcNNs[sensor_name+"_NN"].predict(
                    Experiment.bcNNs[sensor_name+"_SC_X"].transform(rs_ref.transpose())))).transpose())
            init += Experiment.sensor_nobs[i] * len(self.wvl_emulation) * nsims_per_obs
        return reflectance_bands


######New#############
class WhittakerFilter():
    def __init__(self, wvl_obs,lmbda, wvl_required):
        def speyediff(N, d, format='csc'):
            """
            (utility function)
            Construct a d-th order sparse difference matrix based on
            an initial N x N identity matrix
            Final matrix (N-d) x N
            """
            assert not (d < 0), "d must be non negative"
            shape = (N - d, N)
            diagonals = np.zeros(2 * d + 1)
            diagonals[d] = 1.
            for i in range(d):
                diff = diagonals[:-1] - diagonals[1:]
                diagonals = diff
            offsets = np.arange(d + 1)
            spmat = sparse.diags(diagonals, offsets, shape, format=format)
            return spmat
        series=np.arange(400,2501,1)
        self.w = np.isin(series,wvl_obs).astype(float)
        self.mapping = np.isin(series,wvl_required)
        m = len(series)
        E = sp.sparse.diags(self.w, 0, shape=(m, m), format="coo")  # format is coo
        D = speyediff(m, 2, format='csc')
        self.coefmat = E + lmbda * D.conj().T.dot(D)
        # self.C = np.linalg.cholesky(coefmat.toarray()).transpose()
        ###
        # m = len(series)
        # d1 = -1 * np.ones((m), dtype='d')
        # d2 = 3 * np.ones((m), dtype='d')
        # d3 = -3 * np.ones((m), dtype='d')
        # d4 = np.ones((m), dtype='d')
        # D = sp.sparse.diags([d1, d2, d3, d4], [0, 1, 2, 3], shape=(m - 3, m), format="coo")
        # W = sp.sparse.diags(self.w, 0, shape=(m, m), format="coo")
        # t = W + lmbda * (D.transpose()).dot(D)
        # self.C = np.linalg.cholesky(t.toarray()).transpose()
        # print("aa")
    def apply(self,observed):
        interpolated=np.zeros((np.shape(observed)[1],np.sum(self.mapping)))
        for i in np.arange(0,np.shape(observed)[1]):
            obs_mapped=np.zeros_like(self.w)
            obs_mapped[self.w.astype(bool)]=observed[:,i]
            z = splu(self.coefmat).solve(obs_mapped)
            interpolated[i,:]=z[self.mapping]
        return interpolated.transpose()


######New############
class SpectralModel():
    #The spectal model class contains two types of models: a leaf reflectance emulator (PROSPECT) and a soil reflectance model
    #This class creates the mapping of input and output variables to the RTMModelSolution objects and contain the functions to run both models
    def __init__(self, name, param_sequence,output_variables, canopy_model_sequence, Exp, wvl, coef):
        self.coef = coef
        self.name=name
        if self.name == 'emulator_PROSPECT40':
            myshelf = shelve.open('emulators/'+self.name)
            self.emulator_prospect = myshelf["NN"]
            self.input_scaler = myshelf['scaler_input']
            self.output_scaler = myshelf['scaler_output']
            myshelf.close()
        self.param_sequence = param_sequence
        self.used_wvl = np.isin(np.arange(400, 2501, 1), wvl)
        self.ModelParam = Exp.parameters_Frame.loc[Exp.parameters_Frame['model'] == name]
        self.output_positions=[]
        for i in np.arange(0,len(output_variables)):
            self.output_positions.append(np.where(canopy_model_sequence == output_variables[i])[0][0])
        self.ntimes=[]
        self.unique_offset = []
        self.unique_times = []
        self.inverse_rows = []
        self.ntiles = []
        self.nrepeat = []
        for level in np.arange(0, Exp.nlevels):
            npar = Exp.parameters_value[level].shape[1]
            nsims = 1
            activ = False
            ntimes = np.zeros(np.max(self.ModelParam['level'].values) + 1, dtype=int)
            ngroups = np.zeros(np.max(self.ModelParam['level'].values) + 1, dtype=int)
            param_names = []
            # unique_times=np.zeros(np.max(LeafParam['level'].values)+1, dtype=int)
            for level_sim in np.arange(0, np.max(self.ModelParam['level'].values) + 1):
                ngroups_level = np.shape(Exp.parameters_value[level_sim])[2]
                ngroups[level_sim] = ngroups_level
                if activ:
                    param_names = np.concatenate(
                        (param_names, self.ModelParam.index.values[self.ModelParam['level'].values == level_sim]))
                else:
                    param_names = self.ModelParam.index.values[self.ModelParam['level'].values == level_sim]
                    activ = True
                if level_sim == level:
                    nc = 2 * np.sum(self.ModelParam['level'].values == level) * ngroups_level
                    if nc > 0:
                        #    nsims *= nc
                        offset = np.vstack((Exp.deltaGrad * np.identity(npar)[:, np.asarray(
                            self.ModelParam['number_in_level'].values[self.ModelParam['level'] == level], dtype=int)],
                                            -1*Exp.deltaGrad * np.identity(npar)[:, np.asarray(
                                                self.ModelParam['number_in_level'].values[self.ModelParam['level'] == level],
                                                dtype=int)]))
                        unique_sims, inverse = np.unique(offset, return_inverse=True, axis=0)
                        ntimes[level_sim] = ngroups_level
                        unique_offset = np.tile(unique_sims, (ngroups_level, 1))
                        inverse_row = inverse.copy()
                        unique_times = np.shape(unique_sims)[0]
                        for i in np.arange(1, ngroups_level):
                            inverse_row = np.hstack((inverse_row, inverse + i * len(unique_sims)))
                    else:
                        ntimes[level_sim] = 1
                        unique_offset = 0
                        inverse_row = np.asarray([0], dtype=int)
                        inverse = [0]
                        unique_sims = [0]
                        unique_times = 1
                # elif level_sim < level:
                else:
                    # nsims *= ngroups_level
                    ntimes[level_sim] = max((ngroups_level, 1))
                    if 'inverse' in locals():
                        for i in np.arange(1, ngroups_level):
                            inverse_row = np.hstack((inverse_row, inverse + i * len(unique_sims)))
                    else:
                        inverse_row = np.asarray([0], dtype=int)
                        unique_times = 1
            param_order = np.zeros(len(param_names), dtype=int)
            p = 0
            for i in param_sequence:
                param_order[p] = np.where(param_names == i)[0]
                p = p + 1
            ntiles = int((npar*2*Exp.nobs) / len(inverse_row))
            self.ntimes.append(ntimes)
            self.unique_offset.append(unique_offset)
            self.unique_times.append(unique_times)
            self.inverse_rows.append(inverse_row)
            if 'inverse' in locals():
                del inverse
            if np.min(self.ModelParam['level'].values) > level:
                self.nrepeat.append(ntiles)
                self.ntiles.append(1)
            else:
                self.ntiles.append(ntiles)
                self.nrepeat.append(1)
        # level = nlevels when computing refelectance with original values, no gradients
        self.param_order=param_order
        self.ntimes.append(self.ntimes[int(np.max(self.ModelParam['level'].values))])

        self.ntiles.append(self.ntiles[int(np.max(self.ModelParam['level'].values))])
        self.nrepeat.append(self.nrepeat[int(np.max(self.ModelParam['level'].values))])
        self.unique_offset.append(0)
        self.unique_times.append(1)
        self.inverse_rows.append(np.asarray(np.unique(Exp.hierarchical_groups[:,int(np.max(self.ModelParam['level'].values))]),dtype=int))
    def runSpectralModel(self, level,canopyInput, parameters_value):  #########New
        for ilevel in np.arange(0, np.max(self.ModelParam['level'].values) + 1):
            if ilevel < np.max(self.ModelParam['level'].values):
                ntimes_level = self.ntimes[level][ilevel + 1]
            else:
                ntimes_level = 1
            if level == ilevel:
                offset = self.unique_offset[level]
            else:
                offset = 0
            mat = np.tile(np.repeat(np.transpose(parameters_value[ilevel][
                                                 np.asarray(
                                                     self.ModelParam['number_in_level'].values[self.ModelParam['level'] == ilevel],
                                                     dtype=int), :]), self.unique_times[level], axis=0) + offset, (ntimes_level, 1))
            if ilevel == 0:
                parameters = mat.copy()
            else:
                parameters = np.hstack((parameters, mat))

        out_sim = np.zeros((len(self.output_positions),np.shape(parameters)[0], np.sum(self.used_wvl)))
        if self.name == 'emulator_PROSPECT40': #This is to run PROSPECT emulator
            l_ref_tra = self.output_scaler.inverse_transform((self.emulator_prospect.predict(self.input_scaler.transform(parameters[:,self.param_order]))))
            out_sim[0,:,:] = l_ref_tra[:,0: 40]
            out_sim[1,:,:] =  l_ref_tra[:,40:]
        elif self.name == 'soilspec': #This is to run the soil reflectance model
            for simulation in np.arange(0, np.shape(parameters)[0]):
                out_sim[0, simulation,:] = np.dot(self.coef,np.reshape(np.asarray((parameters[simulation, self.param_order[0]],
                                          parameters[simulation, self.param_order[1]],
                                          parameters[simulation, self.param_order[2]],
                                          parameters[simulation, self.param_order[3]])),(-1,1)))[:,0]
        for i in np.arange(0,len(self.output_positions)):
            if self.ntiles[level] > 1 or self.nrepeat[level] > 1:
                canopyInput[:,self.output_positions[i]] = np.tile(np.repeat(out_sim[i,self.inverse_rows[level],:],self.nrepeat[level],axis=0).flatten(),self.ntiles[level])
            else:
                canopyInput[:, self.output_positions[i]] = out_sim[i,self.inverse_rows[level], :].flatten()
        return canopyInput

class CanopyModel():
    def __init__(self, Exp):
        #Creates the canopy reflectance model emulator from the information contained in the Experiment file
        self.name= Exp.canopy_model
        myshelf = shelve.open('emulators/'+self.name)
        self.emulator=myshelf['NN']
        self.input_scaler=myshelf['scaler_input']
        self.output_scaler = myshelf['scaler_output']
        myshelf.close()
        self.sequence = np.asarray(['LAI', 'ALA', 'sun_zenith', 'view_zenith', 'phi_sun_view', 'hotspot', 'leafr', 'leaft', 'soilr','fdir']) #This is the sequence for SAILH, update for other models
        CanopyParam = Exp.parameters_Frame.loc[Exp.parameters_Frame['model'] == Exp.canopy_model]
        self.param_mat=[]
        self.offset=[]
        self.mapping=[]
        self.tiles=[]
        self.ncases_per_obs=[]
        for level in np.arange(0, Exp.nlevels):
            npar = Exp.parameters_value[level].shape[1]
            ngroups = np.shape(Exp.parameters_value[level])[2]
            level_param=[]
            level_offset=[]
            level_mapping=[]
            level_repeat=[]
            for level_sim in np.arange(0, Exp.nlevels + 1):
                if level_sim == Exp.nlevels:
                    npar_level = Exp.parameters_value[level_sim].shape[1]
                    ngroups_level = np.shape(Exp.parameters_value[level_sim])[0]
                else:
                    npar_level = Exp.parameters_value[level_sim].shape[1]
                    ngroups_level = np.shape(Exp.parameters_value[level_sim])[2]
                Modelparam = np.zeros((len(self.sequence), npar_level))
                param_mat = np.repeat(np.identity(ngroups_level), 2 * npar, axis=1)
                if level_sim == level:
                    offset = np.tile(np.hstack((np.identity(npar) * Exp.deltaGrad,np.identity(npar) * -1*Exp.deltaGrad)), (1, ngroups))
                else:
                    offset = 0
                for param in CanopyParam.index.values[CanopyParam['level'].values == level_sim]:
                    Modelparam[np.where(self.sequence == param)[0],
                               CanopyParam['number_in_level'].values[CanopyParam.index.values == param][0]] = 1

                level_param.append(param_mat)
                level_offset.append(offset)
                level_mapping.append(Modelparam)
                level_repeat.append(int((Exp.nobs/ngroups_level)))#*len(wvl))
            self.param_mat.append(level_param)
            self.offset.append(level_offset)
            self.mapping.append(level_mapping)
            self.tiles.append(level_repeat)
            self.ncases_per_obs.append(npar*2)
        # Level nlevels is to compute reflectance for the original values, without gradient
        self.offset.append([0]*(Exp.nlevels+1))
        self.param_mat.append([1] * (Exp.nlevels + 1))
        self.tiles.append(self.tiles[Exp.nlevels-1])
        self.mapping.append(self.mapping[Exp.nlevels - 1])

    def prepareInput(self, level, parameters_value, Exp):
        for level_sim in np.arange(Exp.nlevels+1):
            if level_sim == Exp.nlevels:
                level_values = np.transpose(parameters_value[level_sim][:, :])
            else:
                level_values = parameters_value[level_sim][:, :]
            gradient = np.dot(level_values,self.param_mat[level][level_sim]) + self.offset[level][level_sim]
            mapped = np.dot(self.mapping[level][level_sim],gradient)
            if level_sim==0:
                model_input = np.repeat(np.tile(np.transpose(mapped),(self.tiles[level][level_sim],1)),len(Exp.RTMModelSetUp.wvl_emulation), axis=0)
            else:
                model_input += np.repeat(np.tile(np.transpose(mapped),(self.tiles[level][level_sim],1)),len(Exp.RTMModelSetUp.wvl_emulation), axis=0)
        return model_input

    def Run(self,input_matrix):
        reflectance=(self.output_scaler.inverse_transform((self.emulator.predict(self.input_scaler.transform(input_matrix))).reshape(-1,1))).flatten()
        return reflectance


